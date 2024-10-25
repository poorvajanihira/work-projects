import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook and select the sheet
wb = xl.load_workbook('Uprio Usage sheet.xlsx')
sheet = wb['Daily Usage']

# Initialize a variable to accumulate the total API count
total_api_count = 0

# Iterate through the rows, starting from row 2 to the last row
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 2)  # Assuming the API counts are in the 2nd column
    
    # Check if the cell has a value and is a number
    if cell.value is not None and isinstance(cell.value, (int, float)):
        total_api_count += cell.value  # Accumulate the total count

# Write the total API count to a new cell
total_api_cell = sheet.cell(sheet.max_row + 1, 2)  # Place the total below the last row
total_api_cell.value = total_api_count

values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=2,
            max_col=2)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
# Save the workbook with the updated data
wb.save('updated_sheet.xlsx')
